import openpyxl
from openpyxl import load_workbook
from copy import copy
import string
import statistics
from _pydecimal import Decimal, ROUND_HALF_UP
from openpyxl.styles.borders import Border, Side

#-----add-----
import tkinter as tk
import tkinter.ttk as ttk
import tkinter.font as tkfont
from tkinter import StringVar, filedialog
from tkinter.filedialog import askdirectory
import sys
import os

choose_mode=0
# 組內差分按鈕
def dif_file():
    global input_path
    global output_path
    global choose_mode
    # 讀入檔案
    input_path = filedialog.askopenfilename()

    # 輸出檔案位置
    output_path = input_path.strip(input_path.split("/")[-1])
    output_path = output_path.replace("/","\\") + "output.xlsx"
    choose_mode = 1
    main()

# 組間標準差按鈕
def SD_file():
    global input_path
    global output_path
    global choose_mode

    input_path = filedialog.askopenfilename()

    output_path = input_path.strip(input_path.split("/")[-1])
    output_path = output_path.replace("/","\\") + "output.xlsx"

    choose_mode = 2
    main()

def reset():
    root.destroy()
    restart()

def restart():
    python = sys.executable
    os.execl(python, python, * sys.argv)

step = 0
judge_type=0
grade_type=0
grade_from=0
grade_position=0
counter_for_judge = 0
grade_type_counter = 0
final_grade=[]
base_grade=0
grade_power=0
SD_mode=0
def content_print(event):
    global SD_mode
    global judge_type,grade_type
    global step,counter_for_judge,grade_type_counter
    global all_student_grade,num_teacher_per_group,grade_from,grade_position
    global gap, threshold
    global base_grade,grade_power
    global output_position
    global all_student_mean,all_student_std
    global teacher_grade,teacher_new_grade,student_new_grade,final_grade

    if choose_mode == 0:
       content_back.insert(tk.END,"請先點擊上方按鈕選擇模式")

    # 每一步輸出都會給到下一步 ex.step1輸出完後使用者輸入值將傳進step2，counter同理
    # 組內
    elif choose_mode == 1:
        if step == 0:

            content_back.insert(tk.END,content_value.get())

            input_int = int(content_value.get())
            if input_int == 1 or input_int == 2 or input_int == 3 or input_int == 4 or input_int == 5:
                step = 1
                judge_type = input_int
                if judge_type == 1:
                    content_back.insert(tk.END,"差幾分?")
                elif judge_type == 2:
                    content_back.insert(tk.END,"幾分到幾分是佳？ex:66 80 (請輸入兩個數字):")
                elif judge_type == 3:
                    content_back.insert(tk.END,"委員之間平均評分差 ? 分")
                elif judge_type == 4:
                    content_back.insert(tk.END,"差幾分?")
                elif judge_type == 5:
                    content_back.insert(tk.END,"總成績在哪一行？")
            else:
                content_back.insert(tk.END,"***輸入錯誤，請重新輸入***")

        elif step == 1:
            
            if judge_type == 1:
                #input_student_grade的輸出輸入搬到這邊
                content_back.insert(tk.END,content_value.get())

                if counter_for_judge == 0:

                    threshold = int(content_value.get())

                    content_back.insert(tk.END,"1. 分行複數成績 2.單行複數成績 3.複數個sheet")

                    counter_for_judge = 1

                elif counter_for_judge == 1:

                    input_int = int(content_value.get())

                    if input_int == 1 or input_int == 2 or input_int == 3:

                        grade_type = int(content_value.get())
                        if grade_type == 1:
                            content_back.insert(tk.END,"輸入開始行號與結束行號 ex:A C")
                            content_back.insert(tk.END,"開始行號:")
                        elif grade_type == 2:
                            content_back.insert(tk.END,"成績在哪一行?")
                        elif grade_type == 3:
                            content_back.insert(tk.END,"成績在哪一行?")

                        counter_for_judge = 2

                    else:
                        content_back.insert(tk.END,"***輸入錯誤，請重新輸入***")

                elif counter_for_judge == 2:

                        if grade_type == 1:
                            if grade_type_counter == 0:
                                grade_from = ord(content_value.get())-ord("A")

                                content_back.insert(tk.END,"結束行號:")

                                grade_type_counter = 1

                            elif grade_type_counter == 1:
                                grade_end = ord(content_value.get())-ord("A")
                                grade_type_multiple(grade_from,grade_end)
                                
                                counter_for_judge = 3

                        elif grade_type == 2:
                            grade_where = ord(content_value.get())-ord("A")
                            grade_type_single(grade_where)
                            
                            counter_for_judge = 3

                        elif grade_type == 3:
                            if grade_type_counter == 0:
                                grade_position = ord(content_value.get())-ord("A")

                                content_back.insert(tk.END,"每一組多少位評分委員?")

                                grade_type_counter=1
                            elif grade_type_counter == 1:
                                num_teacher_per_group = int(content_value.get())
                                multiple_sheet(grade_position)

                                counter_for_judge = 3

                #全部資料輸入完成 運算結果
                if counter_for_judge == 3:

                    content_back.insert(tk.END,"-------------------------------------------")
                    content_back.insert(tk.END,"總共" + str(len(all_student_grade)) + "位學生")
                    content_back.insert(tk.END,"請確認第一位學生的成績:")
                    for i in range(len(all_student_grade[0])):
                        content_back.insert(tk.END,str(all_student_grade[0][i]))
                    content_back.insert(tk.END,"請確認最後一位學生的成績:")
                    for i in range(len(all_student_grade[-1])):
                        content_back.insert(tk.END,str(all_student_grade[-1][i]))
                    content_back.insert(tk.END,"不正確的話請請按下重置或請確認表單是否符合要求")
                    content_back.insert(tk.END,"-------------------------------------------")
                    
                    problem_students_data = point_diff(threshold,grade_type)

                    if len(problem_students_data) == 0:
                        content_back.insert(tk.END,"沒有任何差分異常")
                    else:
                        content_back.insert(tk.END,"已生成檔案到" + output_path)
                        output_excel(problem_students_data)

            elif judge_type == 2:
                global good_from,good_end,great_from,great_end
                content_back.insert(tk.END,content_value.get())
                if counter_for_judge == 0:

                    good_from,good_end = map(float,content_value.get().split())

                    content_back.insert(tk.END,"幾分到幾分是優？ex:81 90 (請輸入兩個數字):")

                    counter_for_judge = 1

                elif counter_for_judge == 1:
                    great_from,great_end = map(float,content_value.get().split())

                    content_back.insert(tk.END,"1. 分行複數成績 2.單行複數成績 3.複數個sheet")

                    counter_for_judge = 2

                elif counter_for_judge == 2:

                    input_int = int(content_value.get())

                    if input_int == 1 or input_int == 2 or input_int == 3:

                        grade_type = int(content_value.get())
                        if grade_type == 1:
                            content_back.insert(tk.END,"輸入開始行號與結束行號 ex:A C")
                            content_back.insert(tk.END,"開始行號:")
                        elif grade_type == 2:
                            content_back.insert(tk.END,"成績在哪一行?")
                        elif grade_type == 3:
                            content_back.insert(tk.END,"成績在哪一行?")

                        counter_for_judge = 3
                    else:
                        content_back.insert(tk.END,"***輸入錯誤，請重新輸入***")
                
                elif counter_for_judge == 3:

                    if grade_type == 1:
                        if grade_type_counter == 0:
                            grade_from = ord(content_value.get())-ord("A")

                            content_back.insert(tk.END,"結束行號:")

                            grade_type_counter = 1
                        elif grade_type_counter == 1:
                            grade_end = ord(content_value.get())-ord("A")
                            grade_type_multiple(grade_from,grade_end) #不太懂甚麼叫這邊要偵測error
                            counter_for_judge = 4

                    elif grade_type == 2:
                        grade_where = ord(content_value.get())-ord("A") #不太懂甚麼叫這邊要偵測error
                        grade_type_single(grade_where)
                        counter_for_judge = 4

                    elif grade_type == 3:
                        if grade_type_counter == 0:
                            grade_position = ord(content_value.get())-ord("A")

                            content_back.insert(tk.END,"每一組多少位評分委員?")

                            grade_type_counter=1
                        elif grade_type_counter == 1:
                            num_teacher_per_group = int(content_value.get())
                            multiple_sheet(grade_position)
                            counter_for_judge = 4

                if counter_for_judge == 4:

                    content_back.insert(tk.END,"-------------------------------------------")
                    content_back.insert(tk.END,"總共" + str(len(all_student_grade)) + "位學生")
                    content_back.insert(tk.END,"請確認第一位學生的成績:")
                    for i in range(len(all_student_grade[0])):
                        content_back.insert(tk.END,str(all_student_grade[0][i]))
                    content_back.insert(tk.END,"請確認最後一位學生的成績:")
                    for i in range(len(all_student_grade[-1])):
                        content_back.insert(tk.END,str(all_student_grade[-1][i]))
                    content_back.insert(tk.END,"不正確的話請請按下重置或請確認表單是否符合要求")
                    content_back.insert(tk.END,"-------------------------------------------")

                    problem_students_data = level_diff(good_from,good_end,great_from,great_end,grade_type)

                    if len(problem_students_data) == 0:
                        content_back.insert(tk.END,"沒有任何差分異常")
                    else:
                        content_back.insert(tk.END,"已生成檔案到" + output_path)
                        output_excel(problem_students_data)


            elif judge_type == 3:

                content_back.insert(tk.END,content_value.get())
                if counter_for_judge == 0:

                    gap = float(content_value.get())

                    content_back.insert(tk.END,"同一學生，委員之間平均差距 ? 分")

                    counter_for_judge = 1

                elif counter_for_judge == 1:

                    threshold = float(content_value.get())               

                    content_back.insert(tk.END,"1. 分行複數成績 2.單行複數成績 3.複數個sheet")

                    counter_for_judge = 2

                elif counter_for_judge == 2:

                    input_int = int(content_value.get())
                    if input_int == 1 or input_int == 2 or input_int == 3:

                        grade_type = int(content_value.get())
                        if grade_type == 1:
                            content_back.insert(tk.END,"輸入開始行號與結束行號 ex:A C")
                            content_back.insert(tk.END,"開始行號:")
                        elif grade_type == 2:
                            content_back.insert(tk.END,"成績在哪一行?")
                        elif grade_type == 3:
                            content_back.insert(tk.END,"成績在哪一行?")

                        counter_for_judge = 3
                    else:
                        content_back.insert(tk.END,"***輸入錯誤，請重新輸入***")

                elif counter_for_judge == 3:

                    if grade_type == 1:
                        if grade_type_counter == 0:
                            grade_from = ord(content_value.get())-ord("A")
                            
                            content_back.insert(tk.END,"結束行號:")

                            grade_type_counter = 1
                        elif grade_type_counter == 1:
                            grade_end = ord(content_value.get())-ord("A")
                            grade_type_multiple(grade_from,grade_end) #不太懂甚麼叫這邊要偵測error
                            counter_for_judge = 4

                    elif grade_type == 2:
                        grade_where = ord(content_value.get())-ord("A") #不太懂甚麼叫這邊要偵測error
                        grade_type_single(grade_where)
                        counter_for_judge = 4

                    elif grade_type == 3:
                        if grade_type_counter == 0:
                            grade_position = ord(content_value.get())-ord("A")
                            
                            content_back.insert(tk.END,"每一組多少位評分委員?")

                            grade_type_counter=1
                        elif grade_type_counter == 1:
                            num_teacher_per_group = int(content_value.get())
                            multiple_sheet(grade_position)
                            counter_for_judge = 4

                if counter_for_judge == 4:

                    content_back.insert(tk.END,"-------------------------------------------")
                    content_back.insert(tk.END,"總共" + str(len(all_student_grade)) + "位學生")
                    content_back.insert(tk.END,"請確認第一位學生的成績:")
                    for i in range(len(all_student_grade[0])):
                        content_back.insert(tk.END,str(all_student_grade[0][i]))
                    content_back.insert(tk.END,"請確認最後一位學生的成績:")
                    for i in range(len(all_student_grade[-1])):
                        content_back.insert(tk.END,str(all_student_grade[-1][i]))
                    content_back.insert(tk.END,"不正確的話請請按下重置或請確認表單是否符合要求")
                    content_back.insert(tk.END,"-------------------------------------------")

                    all_professor_mean_grade = professor_mean()

                    for i,professor_mean_grade in enumerate(all_professor_mean_grade):
                        content_back.insert(tk.END,"第" + 
                        str(i+1) + 
                        "位委員打分的平均分數為: " + 
                        str(professor_mean_grade) + " 分")
                    if max(all_professor_mean_grade) - min(all_professor_mean_grade) >= gap:
                        content_back.insert(tk.END,"第" + 
                            str(all_professor_mean_grade.index(min(all_professor_mean_grade))+1) + 
                            "位委員與第" + 
                            str(all_professor_mean_grade.index(max(all_professor_mean_grade))+1) + 
                            "位委員之間的評分大於 " + str(gap) + "分\n")
                    else:
                        content_back.insert(tk.END,"教授評分平均沒有異常")
                        problem_students_data = point_diff(threshold,grade_type)

                        if len(problem_students_data) == 0:
                            content_back.insert(tk.END,"沒有任何差分異常")
                        else:
                            content_back.insert(tk.END,"已生成檔案到" + output_path)
                            output_excel(problem_students_data)

            elif judge_type == 4:
                content_back.insert(tk.END,content_value.get())
                if counter_for_judge == 0:


                    threshold = int(content_value.get())

                    content_back.insert(tk.END,"1. 分行複數成績 2.單行複數成績 3.複數個sheet")

                    counter_for_judge = 1

                elif counter_for_judge == 1:

                    input_int = int(content_value.get())
                    if input_int == 1 or input_int == 2 or input_int == 3:

                        grade_type = int(content_value.get())
                        if grade_type == 1:
                            content_back.insert(tk.END,"輸入開始行號與結束行號 ex:A C")
                            content_back.insert(tk.END,"開始行號:")
                        elif grade_type == 2:
                            content_back.insert(tk.END,"成績在哪一行?")
                        elif grade_type == 3:
                            content_back.insert(tk.END,"成績在哪一行?")

                        counter_for_judge = 2
                    else:
                        content_back.insert(tk.END,"***輸入錯誤，請重新輸入***")

                elif counter_for_judge == 2:

                    if grade_type == 1:
                        if grade_type_counter == 0:
                            grade_from = ord(content_value.get())-ord("A")
                            
                            content_back.insert(tk.END,"結束行號:")

                            grade_type_counter = 1
                        elif grade_type_counter == 1:
                            grade_end = ord(content_value.get())-ord("A")
                            grade_type_multiple(grade_from,grade_end) #不太懂甚麼叫這邊要偵測error
                            counter_for_judge = 3

                    elif grade_type == 2:
                        grade_where = ord(content_value.get())-ord("A") #不太懂甚麼叫這邊要偵測error
                        grade_type_single(grade_where)
                        counter_for_judge = 3

                    elif grade_type == 3:
                        if grade_type_counter == 0:
                            grade_position = ord(content_value.get())-ord("A")
                            
                            content_back.insert(tk.END,"每一組多少位評分委員?")

                            grade_type_counter=1
                        elif grade_type_counter == 1:
                            num_teacher_per_group = int(content_value.get())
                            multiple_sheet(grade_position)
                            counter_for_judge = 3

                #全部資料輸入完成 運算結果
                if counter_for_judge == 3:

                    content_back.insert(tk.END,"-------------------------------------------")
                    content_back.insert(tk.END,"總共" + str(len(all_student_grade)) + "位學生")
                    content_back.insert(tk.END,"請確認第一位學生的成績:")
                    for i in range(len(all_student_grade[0])):
                        content_back.insert(tk.END,str(all_student_grade[0][i]))
                    content_back.insert(tk.END,"請確認最後一位學生的成績:")
                    for i in range(len(all_student_grade[-1])):
                        content_back.insert(tk.END,str(all_student_grade[-1][i]))
                    content_back.insert(tk.END,"不正確的話請按下重置或請確認表單是否符合要求")
                    content_back.insert(tk.END,"-------------------------------------------")

                    problem_students_data = far_from_mean(threshold,grade_type)

                    if len(problem_students_data) == 0:
                        content_back.insert(tk.END,"沒有任何差分異常")
                    else:
                        content_back.insert(tk.END,"已生成檔案到" + output_path)
                        output_excel(problem_students_data)
            elif judge_type == 5:
                grade_position = ord(content_value.get())-ord('A')
                problem_students_data = find_same_grade(grade_position)
                if len(problem_students_data) == 0:
                    content_back.insert(tk.END,"沒有人有相同成績")
                else:
                    output_excel(problem_students_data)

        content_value.set("") #清空輸入框
    # 標準化    
    else:
        # 先判斷是否是社政資管
        if SD_mode == 0:
            content_back.insert(tk.END,content_value.get())
            SD_mode = int(content_value.get())
            content_back.insert(tk.END,"成績在哪一行？")
        # 社政 or 資管
        elif SD_mode == 1 or SD_mode == 2:
            content_back.insert(tk.END,content_value.get())
            if step == 0:
                grade_position = ord(content_value.get())-ord('A')
                for ws in original_wb.worksheets:
                    all_sheet_name.append(ws.title)
                content_back.insert(tk.END,"每一組多少位評分委員？")
                step = 1

            elif step == 1:
                num_teacher_per_group = int(content_value.get())
                content_back.insert(tk.END,"所需要設定的基礎分數值(【A+B*學生分數】中的【A】)")
                step = 2

            elif step == 2:
                base_grade = int(content_value.get())
                content_back.insert(tk.END,"所需要設定的倍率(【A+B*學生分數】中的【B】)")
                step = 3

            elif step == 3:
                grade_power = int(content_value.get())
                grade_temp = []
                time=0
                # 加總單個學生的成績
                for sheet_name in all_sheet_name:
                    iter_ws = original_wb[sheet_name]
                    if time == 0:
                        for i,grade in enumerate(list(iter_ws.columns)[grade_position]):
                            if i == 0:
                                continue
                            grade_temp.append(float(grade.value))
                    else:
                        for i,grade in enumerate(list(iter_ws.columns)[grade_position]):
                            if i == 0:
                                continue
                            grade_temp[i-1] += (float(grade.value))
                    time += 1
                # 單個學生的平均成績(初步成績)
                for number_of_student in range(len(grade_temp)):
                    if SD_mode == 1:
                        grade_temp[number_of_student] = float(Decimal(str(grade_temp[number_of_student])).quantize(Decimal('.00'), ROUND_HALF_UP))
                    else:
                        grade_temp[number_of_student] = float(Decimal(str(grade_temp[number_of_student])).quantize(Decimal('.00000'), ROUND_HALF_UP))
                # 該組平均與標準差
                print("前7位學生成績: ",grade_temp[0:6])
                if SD_mode == 1:    #社政
                    print(statistics.mean(grade_temp))
                    all_student_mean = float(Decimal(str(statistics.mean(grade_temp))).quantize(Decimal('.00'), ROUND_HALF_UP))
                    print(statistics.stdev(grade_temp))
                    print(statistics.pstdev(grade_temp))
                    all_student_std = float(Decimal(str(statistics.stdev(grade_temp))).quantize(Decimal('.00'), ROUND_HALF_UP))
                else:               #資管
                    all_student_mean = float(Decimal(str(statistics.mean(grade_temp))).quantize(Decimal('.00000'), ROUND_HALF_UP))
                    all_student_std = float(Decimal(str(statistics.pstdev(grade_temp))).quantize(Decimal('.00000'), ROUND_HALF_UP))
                print("學生平均: ",all_student_mean)
                print("學生標準差: ",all_student_std)
                for number_of_student in range(len(grade_temp)):
                    # 標準化(標準化成績)
                    # personal_finalgrade為標準化後成績
                    if SD_mode == 1:
                        personal_finalgrade_up = float(Decimal(str((grade_temp[number_of_student]-all_student_mean))).quantize(Decimal('.00'), ROUND_HALF_UP))
                        personal_finalgrade = personal_finalgrade_up/all_student_std
                        personal_finalgrade = float(Decimal(str(personal_finalgrade)).quantize(Decimal('.00000'), ROUND_HALF_UP))
                    else:
                        personal_finalgrade_up = float(Decimal(str((grade_temp[number_of_student]-all_student_mean))).quantize(Decimal('.00000'), ROUND_HALF_UP))
                        personal_finalgrade = personal_finalgrade_up/all_student_std
                        personal_finalgrade = float(Decimal(str(personal_finalgrade)).quantize(Decimal('.00000'), ROUND_HALF_UP))
                    # 最終成績(最終書審成績)
                    # final_grade為最終顯示成績
                    if SD_mode == 1:
                        final_grade.append(float(Decimal(str(personal_finalgrade*grade_power+base_grade)).quantize(Decimal('.00'), ROUND_HALF_UP)))
                    else:
                        final_grade.append(float(Decimal(str(personal_finalgrade*grade_power+base_grade)).quantize(Decimal('.00000'), ROUND_HALF_UP)))
                    # 測試用 輸出第一位學生
                    if number_of_student<3:
                        print("第一位同學成績: ",grade_temp[number_of_student])
                        print("第一位同學標準化後成績: ",personal_finalgrade)
                        print("第一位同學最終成績",final_grade[number_of_student])
                final_grade = [final_grade]
                content_back.insert(tk.END,"輸出框格樣式  直到哪一行？")
                step = 4
            elif step == 4:
                output_position = ord(content_value.get())-ord('A')
                content_back.insert(tk.END,"***資料生成中請稍後***")
                output_wb = openpyxl.workbook.Workbook()
                output_wb.remove(output_wb['Sheet'])
                for i in range(1):
                    # content_back.insert(tk.END,"第"+str(i+1)+"個sheet")
                    output_sheet = output_wb.create_sheet(chr(i+ord("A"))+"組")
                    output_sheet.row_dimensions[1].height = original_wb[all_sheet_name[i*num_teacher_per_group]].row_dimensions[0].height
                    #放入title
                    #因為輸出是從1開始算 所以要加1
                    for the_column in range(output_position+1):
                        #設定所有title的寬
                        # content_back.insert(tk.END,chr(the_column+ord("A")),original_wb[all_sheet_name[i*num_teacher_per_group]].column_dimensions[chr(the_column+ord("A"))].width)
                        output_sheet.column_dimensions[chr(the_column+ord("A"))].width = original_wb[all_sheet_name[i*num_teacher_per_group]].column_dimensions[chr(the_column+ord("A"))].width + 5
                        # content_back.insert(tk.END,output_sheet.column_dimensions[chr(the_column+ord("A"))].width)
                        #放title
                        old_cell = list(original_wb[all_sheet_name[i*num_teacher_per_group]].rows)[0][the_column]
                        new_cell = output_sheet.cell(row=1,column=the_column+1,value=old_cell.value)
                        new_cell.border = copy(old_cell.border)
                    #新增組間標準化後title
                    output_sheet.column_dimensions[chr(output_position+1+ord("A"))].width = 20
                    std_cell = output_sheet.cell(row=1,column=output_position+2,value="組間標準化後成績")
                    std_cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
                    #放入學生資料
                    for the_row in range(len(final_grade[i])):
                        output_sheet.row_dimensions[2+the_row].height = original_wb[all_sheet_name[i*num_teacher_per_group]].row_dimensions[1+the_row].height
                        for the_column in range(output_position+1):
                            old_cell = list(original_wb[all_sheet_name[i*num_teacher_per_group]].rows)[1+the_row][the_column]
                            new_cell = output_sheet.cell(row=2+the_row,column=the_column+1,value=list(original_wb[all_sheet_name[i*num_teacher_per_group]].rows)[1+the_row][the_column].value)
                            new_cell.font = copy(old_cell.font)
                            new_cell.border = copy(old_cell.border)
                            new_cell.fill = copy(old_cell.fill)
                            new_cell.number_format = copy(old_cell.number_format)
                            new_cell.protection = copy(old_cell.protection)
                            new_cell.alignment = copy(old_cell.alignment)
                        new_cell = output_sheet.cell(row=2+the_row,column=the_column+2,value=final_grade[i][the_row])
                        new_cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
                content_back.insert(tk.END,"已生成檔案到"+output_path)
                output_wb.save(output_path)
        # 其他科系
        elif SD_mode == 3:
            content_back.insert(tk.END,content_value.get())
            if step == 0:
                grade_position = ord(content_value.get())-ord('A')
                for ws in original_wb.worksheets:
                    all_sheet_name.append(ws.title)
                content_back.insert(tk.END,"每一組多少位評分委員？")

                step = 1
            elif step == 1:
                num_teacher_per_group = int(content_value.get())
                
                for sheet_name in all_sheet_name:
                    iter_ws = original_wb[sheet_name]
                    temp = []

                    for i,grade in enumerate(list(iter_ws.columns)[grade_position]):
                        if i == 0:
                            continue
                        temp.append(float(grade.value))
                        all_student_grade.append(float(grade.value))
                    teacher_mean.append(float(Decimal(str(statistics.mean(temp))).quantize(Decimal('.0000'), ROUND_HALF_UP)))
                    teacher_std.append(float(Decimal(str(statistics.stdev(temp))).quantize(Decimal('.00000'), ROUND_HALF_UP)))

                all_student_mean = float(Decimal(str(statistics.mean(all_student_grade))).quantize(Decimal('.0000'), ROUND_HALF_UP))
                all_student_std = float(Decimal(str(statistics.stdev(all_student_grade))).quantize(Decimal('.0000000'), ROUND_HALF_UP))

                for iter_teacher,sheet_name in enumerate(all_sheet_name):
                    iter_ws = original_wb[sheet_name]
                    temp_z = []
                    temp_list = []
                    for i,grade in enumerate(list(iter_ws.columns)[grade_position]):
                        if i == 0:
                            continue
                        z = (float(grade.value)-teacher_mean[iter_teacher])/teacher_std[iter_teacher]
                        temp_z.append(float(Decimal(str((z*all_student_std)+all_student_mean)).quantize(Decimal('.0000'), ROUND_HALF_UP)))
                        temp_list.append(float(grade.value))
                    teacher_grade.append(temp_list)
                    teacher_new_grade.append(temp_z)
                
                for iter_group in range(len(all_sheet_name)//num_teacher_per_group):
                    sheet_grade_temp = []
                    #每一位學生
                    for grade in range(len(teacher_new_grade[iter_group*num_teacher_per_group])):
                        #每一位老師
                        temp = 0
                        for i in range(num_teacher_per_group):
                            temp += teacher_new_grade[iter_group*num_teacher_per_group+i][grade]
                        temp /= num_teacher_per_group
                        temp = float(Decimal(str(temp)).quantize(Decimal('.00'), ROUND_HALF_UP))
                        sheet_grade_temp.append(temp)
                    student_new_grade.append(sheet_grade_temp)

                content_back.insert(tk.END,"輸出框格樣式  直到哪一行？")
                
                step = 2
            elif step == 2:
                output_position = ord(content_value.get())-ord('A')
                content_back.insert(tk.END,"***資料生成中請稍後***")
                output_wb = openpyxl.workbook.Workbook()
                output_wb.remove(output_wb['Sheet'])

                for i in range(len(all_sheet_name)//num_teacher_per_group):
                    content_back.insert(tk.END,"第"+str(i+1)+"個sheet")
                    output_sheet = output_wb.create_sheet(chr(i+ord("A"))+"組")
                    output_sheet.row_dimensions[1].height = original_wb[all_sheet_name[i*num_teacher_per_group]].row_dimensions[0].height
                    #放入title
                    #因為輸出是從1開始算 所以要加1
                    for the_column in range(output_position+1):
                        #設定所有title的寬
                        # content_back.insert(tk.END,chr(the_column+ord("A")),original_wb[all_sheet_name[i*num_teacher_per_group]].column_dimensions[chr(the_column+ord("A"))].width)
                        output_sheet.column_dimensions[chr(the_column+ord("A"))].width = original_wb[all_sheet_name[i*num_teacher_per_group]].column_dimensions[chr(the_column+ord("A"))].width + 5
                        # content_back.insert(tk.END,output_sheet.column_dimensions[chr(the_column+ord("A"))].width)
                        #放title
                        old_cell = list(original_wb[all_sheet_name[i*num_teacher_per_group]].rows)[0][the_column]
                        new_cell = output_sheet.cell(row=1,column=the_column+1,value=old_cell.value)
                        new_cell.border = copy(old_cell.border)
                    #新增組間標準化後title
                    output_sheet.column_dimensions[chr(output_position+1+ord("A"))].width = 20
                    std_cell = output_sheet.cell(row=1,column=output_position+2,value="組間標準化後成績")
                    std_cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
                    #放入學生資料
                    for the_row in range(len(student_new_grade[i])):
                        output_sheet.row_dimensions[2+the_row].height = original_wb[all_sheet_name[i*num_teacher_per_group]].row_dimensions[1+the_row].height
                        for the_column in range(output_position+1):
                            old_cell = list(original_wb[all_sheet_name[i*num_teacher_per_group]].rows)[1+the_row][the_column]
                            new_cell = output_sheet.cell(row=2+the_row,column=the_column+1,value=list(original_wb[all_sheet_name[i*num_teacher_per_group]].rows)[1+the_row][the_column].value)
                            new_cell.font = copy(old_cell.font)
                            new_cell.border = copy(old_cell.border)
                            new_cell.fill = copy(old_cell.fill)
                            new_cell.number_format = copy(old_cell.number_format)
                            new_cell.protection = copy(old_cell.protection)
                            new_cell.alignment = copy(old_cell.alignment)
                        new_cell = output_sheet.cell(row=2+the_row,column=the_column+2,value=student_new_grade[i][the_row])
                        new_cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
                print(student_new_grade)
                content_back.insert(tk.END,"已生成檔案到"+output_path)
                output_wb.save(output_path)
    content_back.see("end")
    content_value.set("")
#-----add-----

# original sheet 是從 0 開始算起
# output sheet 是從 1 開始算起

# 輸入單行成績到all_student_grade
def grade_type_single(grade_where):
    global original_sheet
    global all_student_grade
    for i,cell in enumerate(list(original_sheet.columns)[grade_where]):
        if i == 0 or cell.value == None:
            continue
        all_student_grade.append(list(map(float, cell.value.split('\n'))))

# 輸入多行成績到all_student_grade
def grade_type_multiple(grade_from,grade_end):
    global original_sheet
    global all_student_grade
    for cell in list(original_sheet.rows):
        if str(cell[0].value).isdigit(): 
            temp = []
            for i in range(grade_from,grade_end+1):
                temp.append(float(cell[i].value))
            all_student_grade.append(temp)
        else:
            continue

# 讀多個sheet到all_student_grade
def multiple_sheet(grade_position):
    global original_wb
    global all_student_grade
    global all_sheet_name
    global num_teacher_per_group
    for ws in original_wb.worksheets:
        all_sheet_name.append(ws.title)
    
    teacher_grade = []
    for sheet_name in all_sheet_name:
        iter_ws = original_wb[sheet_name]
        temp_list = []
        for i,grade in enumerate(list(iter_ws.columns)[grade_position]):
            if i == 0:
                continue
            temp_list.append(float(grade.value))

        teacher_grade.append(temp_list)
    for it1 in range(len(teacher_grade)//num_teacher_per_group): # 總共有幾個group
        for it2 in range(len(teacher_grade[it1*num_teacher_per_group])): #那個group總共有幾個學生
            temp2 = []
            for it3 in range(num_teacher_per_group): #每一位學生有幾個成績
                temp2.append(teacher_grade[(it1*num_teacher_per_group)+it3][it2])
            all_student_grade.append(temp2)

# 輸出到excel
def output_excel(problem_students_data):
    global num_title
    global original_sheet
    output_wb = openpyxl.Workbook()
    output_sheet = output_wb['Sheet']
    output_sheet.row_dimensions[1].height = original_sheet.row_dimensions[0].height
    # 放入title
    for the_column in range(num_title):
        # 設定所有title的寬
        output_sheet.column_dimensions[chr(the_column+ord("A"))].width = original_sheet.column_dimensions[chr(the_column+ord("A"))].width + 5
        # 放title
        old_cell = list(original_sheet.rows)[0][the_column]
        new_cell = output_sheet.cell(row=1,column=the_column+1,value=old_cell.value)
        new_cell.border = copy(old_cell.border)
    # 放入學生資料
    for the_row in range(len(problem_students_data)):
        output_sheet.row_dimensions[2+the_row].height = original_sheet.row_dimensions[1+the_row].height
        for the_column in range(num_title):
            old_cell = list(original_sheet.rows)[1][the_column] # 反正我就已經規定第一行是title 第二開始是資料 所以直接用第二行的樣式
            new_cell = output_sheet.cell(row=2+the_row,column=the_column+1,value=problem_students_data[the_row][the_column])
            new_cell.font = copy(old_cell.font)
            new_cell.border = copy(old_cell.border)
            new_cell.fill = copy(old_cell.fill)
            new_cell.number_format = copy(old_cell.number_format)
            new_cell.protection = copy(old_cell.protection)
            new_cell.alignment = copy(old_cell.alignment)

    output_wb.save(output_path) # 這邊到時候要處理

# 1. 單一學生的分差
def point_diff(threshold,grade_type):
    global all_student_grade
    global original_sheet
    global num_title

    global original_wb
    global num_teacher_per_group
    global all_sheet_name

    problem_students_data = []
    # 不是multiple sheet
    if grade_type != 3:
        for i in range(len(all_student_grade)):
            if float((max(all_student_grade[i])-min(all_student_grade[i]))) > threshold:
                temp = []
                for x in range(num_title):
                    temp.append(list(original_sheet.rows)[i+1][x].value)
                problem_students_data.append(temp)

    # 是multiple sheet的分析輸出不一樣
    else:
        for i in range(len(all_sheet_name)//num_teacher_per_group):
            count = 0
            for the_row in range(len(list(original_wb[all_sheet_name[i*num_teacher_per_group]].columns)[0])):
                if the_row == 0:
                    continue
                if float(max(all_student_grade[count])-min(all_student_grade[count])) > threshold:
                    for x in range(num_teacher_per_group):
                        temp = []
                        for the_column in range(num_title):
                            temp.append(list(original_wb[all_sheet_name[(i*num_teacher_per_group)+x]].rows)[count+1][the_column].value)
                        problem_students_data.append(temp)
                count+=1
             
    return problem_students_data

# 2. 相差兩個級距 是不是應該乾脆全部輸入?
def level_diff(good_from,good_end,great_from,great_end,grade_type):
    
    def level_definition(point):
        if point > great_end:
            return 4
        elif point >= great_from and point <= great_end:
            return 3
        elif point >= good_from and point <= good_end:
            return 2
        elif point < good_from:
            return 1
        else:
            # error
            # print("Error")
            content_back.insert(tk.END,"***level_diff ERROR***")
    
    global all_student_grade
    global original_sheet
    global num_title

    global original_wb
    global num_teacher_per_group
    global all_sheet_name

    problem_students_data = []

    if grade_type != 3:
        for i in range(len(all_student_grade)):
            # 小數點4捨5入後 再去看是哪個區間
            if level_definition(round(max(all_student_grade[i])))-level_definition(round(min(all_student_grade[i]))) >= 2:
                temp = []
                for x in range(num_title):
                    temp.append(list(original_sheet.rows)[i+1][x].value)
                problem_students_data.append(temp)

    else:
        for i in range(len(all_sheet_name)//num_teacher_per_group):
            count = 0
            for the_row in range(len(list(original_wb[all_sheet_name[i*num_teacher_per_group]].columns)[0])):
                if the_row == 0:
                    continue
                if level_definition(round(max(all_student_grade[count])))-level_definition(round(min(all_student_grade[count]))) >= 2:
                    for x in range(num_teacher_per_group):
                        temp = []
                        for the_column in range(num_title):
                            temp.append(list(original_wb[all_sheet_name[(i*num_teacher_per_group)+x]].rows)[count+1][the_column].value)
                        problem_students_data.append(temp)
                count+=1

    return problem_students_data

# 3. 輸出教授打分平均
def professor_mean():
    global original_sheet
    global num_title
    global all_student_grade
    num_professor  = len(all_student_grade[0])
    all_professor_grade = [0.0] * num_professor
    for student_grade in all_student_grade:
        for i in range(num_professor):
            all_professor_grade[i] += student_grade[i]

    for i in range(len(all_professor_grade)):
        all_professor_grade[i] = all_professor_grade[i]/float(len(all_student_grade))
    
    return all_professor_grade
     

# 4. 學生差距平均評分 ? 分以上
def far_from_mean(threshold,grade_type):
    global original_sheet
    global num_title
    global all_student_grade
    num_grade = 0
    total_grade = 0
    problem_students_data = []
    for student_grade in all_student_grade:
        num_grade+=len(student_grade)
        for grade in student_grade:
            total_grade+=grade
    mean = total_grade/num_grade
    content_back.insert(tk.END,"平均分數: " + str(mean) + "分")

    if grade_type != 3:
        for grade_iterator,student_grade in enumerate(all_student_grade):
            if statistics.mean(student_grade) >= (mean+threshold) or statistics.mean(student_grade) <= (mean-threshold):
                temp = []
                for i in range(num_title):
                    temp.append(list(original_sheet.rows)[grade_iterator+1][i].value)
                    
                problem_students_data.append(temp)
    else:
        for i in range(len(all_sheet_name)//num_teacher_per_group):
            count = 0
            for the_row in range(len(list(original_wb[all_sheet_name[i*num_teacher_per_group]].columns)[0])):
                if the_row == 0:
                    continue
                if statistics.mean(all_student_grade[count]) >= (mean+threshold) or statistics.mean(all_student_grade[count]) <= (mean-threshold):
                    for x in range(num_teacher_per_group):
                        temp = []
                        for the_column in range(num_title):
                            temp.append(list(original_wb[all_sheet_name[(i*num_teacher_per_group)+x]].rows)[count+1][the_column].value)
                        problem_students_data.append(temp)
                    pass
                count+=1

    return problem_students_data

# 5. 回傳相同成績的學生data 只吃單行單個成績
def find_same_grade(grade_position):
    global original_sheet
    all_student_grade_position = {}
    problem_students_data = []
    for i,cell in enumerate(list(original_sheet.columns)[grade_position]):
        if i == 0 or cell.value == None:
            continue
        if float(cell.value) not in all_student_grade_position:
            all_student_grade_position[float(cell.value)] = list()
        all_student_grade_position[float(cell.value)].append(i-1)
    for grade in all_student_grade_position:
        if len(all_student_grade_position[grade]) > 1:
            for position in all_student_grade_position[grade]:
                temp = []
                for i in range(num_title):
                    temp.append(list(original_sheet.rows)[position+1][i].value)
                problem_students_data.append(temp)
    return problem_students_data

def main():
    global original_wb
    global original_sheet
    global num_title
    global choose_mode
    original_wb = load_workbook(input_path) 
    original_sheet = original_wb.active

    num_title = 0 # initialize
    # 計算title數量
    for cell in list(original_sheet.rows)[0]:
        if cell.value == None:
            break
        else:
            num_title+=1

    if choose_mode == 1:
        content_back.insert(tk.END,"輸入判斷標準")
        content_back.insert(tk.END,"1. 委員之間差距 ? 分")
        content_back.insert(tk.END,"2. 委員之間差距兩級以上")
        content_back.insert(tk.END,"3. 委員之間差距 ? 分 或各委員平均評分差距 ? 分")
        content_back.insert(tk.END,"4. 學生差距平均評分 ? 分以上")
        content_back.insert(tk.END,"5. 找出相同成績的同學")

    else:
        content_back.insert(tk.END,"請選擇模式")
        content_back.insert(tk.END,"1. 社政")
        content_back.insert(tk.END,"2. 資管")
        content_back.insert(tk.END,"3. 其他科系")
    
original_wb = None # 讀新版檔案
original_sheet = None # 原始的sheet
num_title = None # title數量
all_student_grade = [] # 純總成績

# multiple sheet用的
all_sheet_name = []
num_teacher_per_group = 0

# 標準化
#-----add-----
teacher_mean = []
teacher_std = []
all_student_mean = None
all_student_std = None
teacher_grade = []
teacher_new_grade = []
student_new_grade = []
output_position = None

# 視窗
#-----add-----
root = tk.Tk()
root.title('差分檢核系統')
root.geometry('600x700')
root.configure(background='CornflowerBlue')
root.resizable(0,0)

header_font = tkfont.Font(family="標楷體",size=20,weight='bold')
button_font = tkfont.Font(family="標楷體",size=14)
content_font = tkfont.Font(family="Bahnschrift",size=12)

header_label = tk.Label(root, text='差分檢核系統',font=header_font,fg='white')
header_label.configure(background='CornflowerBlue')
header_label.pack(side=tk.TOP,pady=20)

# 上半
upper = tk.Label(root,width=400,height=200)
upper.configure(background='CornflowerBlue')
upper.pack(side=tk.TOP,fill=tk.X)

# 組內差分 按鈕
dif_text = tk.StringVar()
dif_text.set("組內差分")
select_file = tk.Button(upper,
    textvariable = dif_text,
    command = dif_file,
    width = 20,
    height = 4,
    wraplength = 160,
    justify = 'left',
    font=button_font)
select_file.pack(side=tk.LEFT,padx=40)

# 組間標準差 按鈕
SD_text = tk.StringVar()
SD_text.set("組間標準差")
select_file = tk.Button(upper,
    textvariable = SD_text,
    command = SD_file,
    width = 20,
    height = 4,
    wraplength = 160,
    justify = 'left',
    font=button_font)
select_file.pack(side=tk.RIGHT,padx=40)

# 下半
lower = tk.Label(root,width=400,height=200)
lower.configure(background='LightBlue')
lower.pack(side=tk.TOP,fill=tk.X,pady=30)

# 輸出框
content_area = tk.Label(lower,bg='LightBlue',width=68,height=20)
content_area.pack(side=tk.TOP,padx=30,pady=10)

content_yscrollbar = tk.Scrollbar(content_area)
content_yscrollbar.pack(side=tk.RIGHT,fill=tk.Y)

content_xscrollbar = tk.Scrollbar(content_area,orient=tk.HORIZONTAL)
content_xscrollbar.pack(side=tk.BOTTOM,fill=tk.X)

content_back = tk.Listbox(content_area,
    width=68,
    height=12,
    justify='left',
    font=content_font,
    xscrollcommand = content_xscrollbar.set,
    yscrollcommand = content_yscrollbar.set)
content_back.pack(side=tk.LEFT,fill=tk.BOTH)
content_back.insert(tk.END,"<<歡迎使用差分檢核系統>>")

content_yscrollbar.config(command=content_back.yview)
content_xscrollbar.config(command=content_back.xview)

# 輸入框
content_value = tk.StringVar()
content = tk.Entry(lower
    ,textvariable=content_value
    ,width=75
    ,justify='left')
content.bind("<Return>",content_print)
content.pack(side=tk.TOP,pady=20)

reset_button = tk.Button(root,
    text="重置",
    command=reset,
    width=5,
    height=2,
    font=button_font)
reset_button.pack(side=tk.TOP)

root.mainloop()
#-----add-----